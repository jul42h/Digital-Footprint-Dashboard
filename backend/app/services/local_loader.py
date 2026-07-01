"""Parse Shodan Excel into DashboardData (mirrors frontend dataTransformers)."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from app.models import CVEFlatRecord, DashboardData, DashboardStats, SourceCVE, SourceIPRecord

COLUMN_ALIASES: dict[str, list[str]] = {
    "ip": ["ip address", "ip", "ip_address"],
    "organization": ["organization", "org", "company"],
    "country": ["country", "country code", "location country code", "location_country_code"],
    "city": ["city", "location city", "location_city"],
    "asn": ["asn"],
    "hostnames": ["hostnames", "hostname", "host"],
    "operatingSystem": ["operating system", "os", "operating_system"],
    "ports": ["ports", "port"],
    "transport": ["transport"],
    "service": ["service", "services"],
    "product": ["product", "products"],
    "version": ["version", "versions"],
    "cve": ["cve", "cve id", "cve_id"],
    "cveScore": ["cve score", "cvss score", "cvss", "score", "cvss v2", "cvss_v2"],
    "cvssSeverity": ["cvss severity", "severity", "risk level"],
    "publishedDate": ["published date", "published", "published_date", "observed at", "observed_at"],
    "summary": ["summary", "description"],
    "kev": ["kev", "known exploited"],
    "timestamp": ["observed_at", "timestamp", "last seen", "last_seen"],
}


def _norm(key: str) -> str:
    return key.strip().lower().replace("_", " ").replace("-", " ")


def _cell(row: dict, field: str) -> str:
    aliases = COLUMN_ALIASES.get(field, [])
    for key, value in row.items():
        if _norm(str(key)) in aliases and value is not None:
            return str(value).strip()
    return ""


def _split_list(value: str) -> list[str]:
    if not value:
        return []
    return [p.strip() for p in value.replace("|", ",").replace(";", ",").split(",") if p.strip()]


def _parse_ports(value: str) -> list[int]:
    ports: list[int] = []
    for part in _split_list(value):
        for token in part.split():
            digits = "".join(c for c in token if c.isdigit())
            if digits:
                ports.append(int(digits))
    return ports


def _severity(score: float, raw: str) -> str:
    if raw:
        title = raw.strip().title()
        if title in {"Critical", "High", "Medium", "Low", "Informational"}:
            return title
    if score >= 9:
        return "Critical"
    if score >= 7:
        return "High"
    if score >= 4:
        return "Medium"
    if score > 0:
        return "Low"
    return "Informational"


def _rows_from_excel(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows_iter)]
    rows: list[dict] = []
    for values in rows_iter:
        row = {headers[i]: values[i] for i in range(len(headers)) if headers[i]}
        if any(v is not None and str(v).strip() for v in row.values()):
            rows.append(row)
    wb.close()
    return rows


def _build_cve(row: dict) -> SourceCVE | None:
    cve_id = _cell(row, "cve")
    if not cve_id:
        return None
    score_raw = _cell(row, "cveScore")
    score = float(score_raw) if score_raw else 0.0
    sev = _severity(score, _cell(row, "cvssSeverity"))
    kev_raw = _cell(row, "kev").lower()
    return SourceCVE(
        id=cve_id,
        score=score,
        severity=sev,
        publishedDate=_cell(row, "publishedDate"),
        summary=_cell(row, "summary") or None,
        kev=kev_raw in {"true", "yes", "1", "y"},
    )


def _merge_ip(existing: SourceIPRecord, row: dict) -> SourceIPRecord:
    cve = _build_cve(row)
    cves = list(existing.cves)
    if cve and cve.id not in {c.id for c in cves}:
        cves.append(cve)

    ports = list(dict.fromkeys(existing.ports + _parse_ports(_cell(row, "ports"))))
    transport = list(dict.fromkeys(existing.transport + _split_list(_cell(row, "transport"))))
    services = list(dict.fromkeys(existing.services + _split_list(_cell(row, "service"))))
    products = list(dict.fromkeys(existing.products + _split_list(_cell(row, "product"))))
    versions = list(dict.fromkeys(existing.versions + _split_list(_cell(row, "version"))))
    hostnames = list(dict.fromkeys(existing.hostnames + _split_list(_cell(row, "hostnames"))))

    risk_order = ["Critical", "High", "Medium", "Low", "Informational"]
    risk = existing.riskLevel
    for c in cves:
        if risk_order.index(c.severity) < risk_order.index(risk):
            risk = c.severity

    return existing.model_copy(
        update={
            "organization": existing.organization or _cell(row, "organization"),
            "country": existing.country or _cell(row, "country"),
            "city": existing.city or _cell(row, "city") or None,
            "asn": existing.asn or _cell(row, "asn") or None,
            "operatingSystem": existing.operatingSystem or _cell(row, "operatingSystem") or None,
            "hostnames": hostnames,
            "ports": ports,
            "openPorts": ports,
            "transport": transport,
            "services": services,
            "products": products,
            "versions": versions,
            "cves": cves,
            "riskLevel": risk,
            "timestamp": existing.timestamp or _cell(row, "timestamp") or None,
            "lastSeen": _cell(row, "timestamp") or existing.lastSeen,
        }
    )


def transform_rows(rows: list[dict]) -> list[SourceIPRecord]:
    by_ip: dict[str, SourceIPRecord] = {}
    for row in rows:
        ip = _cell(row, "ip")
        if not ip:
            continue
        if ip not in by_ip:
            by_ip[ip] = SourceIPRecord(ip=ip, organization=_cell(row, "organization"), country=_cell(row, "country"))
        by_ip[ip] = _merge_ip(by_ip[ip], row)
    return list(by_ip.values())


def compute_stats(ips: list[SourceIPRecord]) -> DashboardStats:
    all_cves: list[SourceCVE] = []
    orgs: set[str] = set()
    countries: set[str] = set()
    dates: list[str] = []

    for ip in ips:
        if ip.organization:
            orgs.add(ip.organization)
        if ip.country:
            countries.add(ip.country)
        all_cves.extend(ip.cves)

    counts = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0, "Informational": 0}
    scores: list[float] = []
    for cve in all_cves:
        counts[cve.severity] = counts.get(cve.severity, 0) + 1
        scores.append(cve.score)
        if cve.publishedDate:
            dates.append(cve.publishedDate)

    return DashboardStats(
        totalIPs=len(ips),
        totalCVEs=len(all_cves),
        criticalCVEs=counts["Critical"],
        highCVEs=counts["High"],
        mediumCVEs=counts["Medium"],
        lowCVEs=counts["Low"],
        informationalCVEs=counts["Informational"],
        averageCVSS=round(sum(scores) / len(scores), 2) if scores else 0,
        highestCVSS=max(scores) if scores else 0,
        newestVulnerability=max(dates) if dates else None,
        oldestVulnerability=min(dates) if dates else None,
        uniqueOrganizations=len(orgs),
        uniqueCountries=len(countries),
    )


def flatten_cves(ips: list[SourceIPRecord]) -> list[CVEFlatRecord]:
    flat: list[CVEFlatRecord] = []
    for ip in ips:
        port = ip.ports[0] if ip.ports else None
        for cve in ip.cves:
            flat.append(
                CVEFlatRecord(
                    cve=cve,
                    ip=ip.ip,
                    organization=ip.organization,
                    country=ip.country,
                    operatingSystem=ip.operatingSystem,
                    port=port,
                )
            )
    return flat


def load_local_dashboard(path: Path) -> DashboardData:
    rows = _rows_from_excel(path)
    ips = transform_rows(rows)
    return DashboardData(
        ips=ips,
        stats=compute_stats(ips),
        cveRecords=flatten_cves(ips),
        lastUpdated=datetime.now(timezone.utc).isoformat(),
        source="excel",
    )
